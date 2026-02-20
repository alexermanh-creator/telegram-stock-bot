
import os
import telebot
import matplotlib.pyplot as plt
from telebot.types import ReplyKeyboardMarkup
from openpyxl import Workbook, load_workbook
from datetime import datetime
from portfolio import *

TOKEN = os.getenv("TELEGRAM_TOKEN")
bot = telebot.TeleBot(TOKEN)

init_db()

state = {}

def main_menu():
    m = ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("📊 Tài sản", "📜 Lịch sử")
    m.row("➕ Nạp thêm", "➖ Rút ra")
    m.row("✏️ Sửa giao dịch", "❌ Xóa giao dịch")
    m.row("💵 Giá trị hiện tại", "📈 Biểu đồ")
    m.row("📥 Import Excel", "📤 Xuất Excel")
    m.row("📘 Hướng dẫn")
    return m

def category_menu():
    m = ReplyKeyboardMarkup(resize_keyboard=True)
    m.row("🪙 Crypto", "📈 Stock")
    m.row("⬅️ Quay lại")
    return m

@bot.message_handler(commands=['start'])
def start(msg):
    bot.send_message(msg.chat.id, "🤖 Bot Quản Lý Tài Sản PRO (Buttons Only)", reply_markup=main_menu())

# ===== HELP =====
@bot.message_handler(func=lambda m: m.text == "📘 Hướng dẫn")
def help_menu(msg):
    bot.send_message(msg.chat.id,
        "Dùng nút để thao tác.\n"
        "Nạp/Rút/Value → chọn danh mục → nhập số tiền.\n"
        "Sửa/Xóa → xem Lịch sử để lấy ID.",
        reply_markup=main_menu()
    )

# ===== REPORT =====
@bot.message_handler(func=lambda m: m.text == "📊 Tài sản")
def report(msg):
    data, total_value, total_profit, total_percent = get_report(msg.from_user.id)
    text = "📊 TÀI SẢN\n\n"
    for cat, d in data.items():
        name = "Crypto" if cat == "crypto" else "Chứng khoán"
        text += f"{name}\nNạp: {d['deposit']:,.0f}\nRút: {d['withdraw']:,.0f}\nGiá trị: {d['value']:,.0f}\nLãi/Lỗ: {d['profit']:,.0f} ({d['percent']:.2f}%)\n\n"
    text += f"💰 Tổng tài sản: {total_value:,.0f}\n📈 Tổng lãi/lỗ: {total_profit:,.0f} ({total_percent:.2f}%)"
    bot.send_message(msg.chat.id, text, reply_markup=main_menu())

# ===== HISTORY =====
@bot.message_handler(func=lambda m: m.text == "📜 Lịch sử")
def history(msg):
    rows = get_history(msg.from_user.id)
    if not rows:
        bot.send_message(msg.chat.id, "Chưa có dữ liệu", reply_markup=main_menu()); return
    text = "📜 Lịch sử (20 gần nhất)\n\n"
    for tx_id, cat, ttype, amount, date in rows[-20:]:
        icon = "📥" if ttype=="deposit" else "📤"
        text += f"ID:{tx_id} {icon} {cat} {amount:,.0f} | {date}\n"
    bot.send_message(msg.chat.id, text, reply_markup=main_menu())

# ===== FLOWS =====
def start_flow(chat_id, action):
    state[chat_id] = {"action": action}
    bot.send_message(chat_id, "Chọn danh mục:", reply_markup=category_menu())

@bot.message_handler(func=lambda m: m.text == "➕ Nạp thêm")
def flow_deposit(msg): start_flow(msg.chat.id, "deposit")

@bot.message_handler(func=lambda m: m.text == "➖ Rút ra")
def flow_withdraw(msg): start_flow(msg.chat.id, "withdraw")

@bot.message_handler(func=lambda m: m.text == "💵 Giá trị hiện tại")
def flow_value(msg): start_flow(msg.chat.id, "value")

@bot.message_handler(func=lambda m: m.text in ["🪙 Crypto","📈 Stock"])
def choose_cat(msg):
    if msg.chat.id not in state: return
    cat = "crypto" if "Crypto" in msg.text else "stock"
    state[msg.chat.id]["category"] = cat
    bot.send_message(msg.chat.id, "Nhập số tiền:")

@bot.message_handler(func=lambda m: m.chat.id in state and "category" in state.get(m.chat.id, {}))
def input_amount(msg):
    try:
        data = state[msg.chat.id]
        amount = float(msg.text)
        cat = data["category"]
        today = datetime.now().strftime("%Y-%m-%d")
        if data["action"] == "deposit":
            add_transaction(msg.from_user.id, cat, "deposit", amount, today)
            bot.send_message(msg.chat.id, "✅ Đã nạp", reply_markup=main_menu())
        elif data["action"] == "withdraw":
            add_transaction(msg.from_user.id, cat, "withdraw", amount, today)
            bot.send_message(msg.chat.id, "✅ Đã rút", reply_markup=main_menu())
        elif data["action"] == "value":
            set_value(msg.from_user.id, cat, amount)
            bot.send_message(msg.chat.id, "✅ Đã cập nhật giá trị", reply_markup=main_menu())
        state.pop(msg.chat.id, None)
    except:
        bot.send_message(msg.chat.id, "❌ Số không hợp lệ")

@bot.message_handler(func=lambda m: m.text == "⬅️ Quay lại")
def back(msg):
    state.pop(msg.chat.id, None)
    bot.send_message(msg.chat.id, "Menu chính", reply_markup=main_menu())

# ===== EDIT / DELETE =====
@bot.message_handler(func=lambda m: m.text == "✏️ Sửa giao dịch")
def edit_info(msg):
    bot.send_message(msg.chat.id, "Nhập: ID số_tiền yyyy-mm-dd")

@bot.message_handler(func=lambda m: m.text == "❌ Xóa giao dịch")
def del_info(msg):
    bot.send_message(msg.chat.id, "Nhập ID cần xóa")

@bot.message_handler(func=lambda m: m.text and m.text.split()[0].isdigit())
def edit_or_delete(msg):
    parts = msg.text.split()
    try:
        if len(parts)==1:
            delete_transaction(msg.from_user.id, int(parts[0]))
            bot.send_message(msg.chat.id, "✅ Đã xóa", reply_markup=main_menu())
        elif len(parts)==3:
            tx_id=int(parts[0]); amount=float(parts[1]); date=parts[2]
            update_transaction(msg.from_user.id, tx_id, amount, date)
            bot.send_message(msg.chat.id, "✅ Đã sửa", reply_markup=main_menu())
    except:
        bot.send_message(msg.chat.id, "❌ Sai cú pháp")

# ===== CHART =====
@bot.message_handler(func=lambda m: m.text == "📈 Biểu đồ")
def chart(msg):
    rows = get_history(msg.from_user.id)
    if not rows:
        bot.send_message(msg.chat.id, "Chưa có dữ liệu", reply_markup=main_menu()); return
    dates=[]; totals=[]; total=0
    for _,_,ttype,amount,date in rows:
        total += amount if ttype=="deposit" else -amount
        dates.append(str(date)); totals.append(total)
    plt.figure(); plt.plot(dates, totals); plt.xticks(rotation=45); plt.tight_layout()
    fname="chart.png"; plt.savefig(fname); plt.close()
    with open(fname,"rb") as f: bot.send_photo(msg.chat.id,f)
    os.remove(fname)

# ===== IMPORT / EXPORT =====
def to_date_str(val):
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    return str(val)

def to_float(val):
    try: return float(val)
    except: return None

@bot.message_handler(func=lambda m: m.text == "📥 Import Excel")
def import_info(msg):
    bot.send_message(msg.chat.id, "Gửi file Excel (3 bảng Crypto/Stock)")

@bot.message_handler(content_types=['document'])
def handle_doc(msg):
    try:
        fi = bot.get_file(msg.document.file_id)
        data = bot.download_file(fi.file_path)
        fname="import.xlsx"
        with open(fname,"wb") as f: f.write(data)
        wb=load_workbook(fname, data_only=True); ws=wb.active
        count=0
        for row in ws.iter_rows(min_row=6, max_col=4, values_only=True):
            di,ai,do,ao=row
            a=to_float(ai)
            if di and a: add_transaction(msg.from_user.id,"crypto","deposit",a,to_date_str(di)); count+=1
            a=to_float(ao)
            if do and a: add_transaction(msg.from_user.id,"crypto","withdraw",a,to_date_str(do)); count+=1
        for row in ws.iter_rows(min_row=6, min_col=7, max_col=10, values_only=True):
            di,ai,do,ao=row
            a=to_float(ai)
            if di and a: add_transaction(msg.from_user.id,"stock","deposit",a,to_date_str(di)); count+=1
            a=to_float(ao)
            if do and a: add_transaction(msg.from_user.id,"stock","withdraw",a,to_date_str(do)); count+=1
        os.remove(fname)
        bot.send_message(msg.chat.id, f"✅ Import thành công {count} giao dịch", reply_markup=main_menu())
    except Exception as e:
        bot.send_message(msg.chat.id, f"❌ Lỗi import: {e}", reply_markup=main_menu())

@bot.message_handler(func=lambda m: m.text == "📤 Xuất Excel")
def export_excel(msg):
    rows=get_history(msg.from_user.id)
    if not rows:
        bot.send_message(msg.chat.id,"Không có dữ liệu",reply_markup=main_menu()); return
    fname="export.xlsx"; wb=Workbook(); ws=wb.active
    ws.append(["ID","Category","Type","Amount","Date"])
    for r in rows: ws.append(r)
    wb.save(fname)
    with open(fname,"rb") as f: bot.send_document(msg.chat.id,f)
    os.remove(fname)

bot.infinity_polling()
